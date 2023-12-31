import streamlit as st
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px
import json
import ydata_profiling
from openpyxl import load_workbook
from streamlit_pandas_profiling import st_profile_report
st.set_option('deprecation.showPyplotGlobalUse', False)

# Function to load data from different file formats
def load_data(file_path, file_format, sheet_name=None):
    if file_format == 'CSV':
        return pd.read_csv(file_path)
    elif file_format in ['XLS', 'XLSX']:
        if sheet_name is None:
            # Detect sheet names without reading the data
            wb = load_workbook(file_path, read_only=True)
            return wb.sheetnames
        else:
            xls = pd.ExcelFile(file_path)
            return xls.parse(sheet_name)
    elif file_format == 'JSON':
        return pd.read_json(file_path)
    elif file_format == 'TSV':
        return pd.read_csv(file_path, sep='\t')
    else:
        raise ValueError("Unsupported file format. Supported formats are CSV, XLS, XLSX, JSON, and TSV.")

# Function to load inbuilt dataset
def load_inbuilt_dataset(dataset_name):
    dataset_name = dataset_name.lower()
    if dataset_name == 'iris':
        return sns.load_dataset('iris')
    elif dataset_name == 'tips':
        return sns.load_dataset('tips')
    elif dataset_name == 'titanic':
        return sns.load_dataset('titanic')
    else:
        raise ValueError("Unsupported dataset.")

# Function for automated EDA
def automated_eda(data):
    # Summary statistics
    summary = data.describe()

    # Data types
    data_types = data.dtypes

    # Missing values
    missing_values = data.isnull().sum()

    # Exclude non-numeric columns from correlation matrix calculation
    numeric_columns = data.select_dtypes(include=['int64', 'float64']).columns
    correlation_matrix = data[numeric_columns].corr()

    # Distribution plots
    for column in numeric_columns:
        fig, ax = plt.subplots()
        sns.histplot(data[column], kde=True)
        plt.title(f'Distribution of {column}')
        plt.xlabel(column)
        plt.ylabel('Frequency')
        st.pyplot(fig)

    # Correlation heatmap
    fig, ax = plt.subplots(figsize=(10, 8))
    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', linewidths=0.5)
    plt.title('Correlation Heatmap')
    st.pyplot(fig)

    # Pairwise scatter plots (for numeric columns)
    if len(numeric_columns) >= 2:
        pair_plot = sns.pairplot(data=data, vars=numeric_columns)
        pair_plot.fig.suptitle('Pairwise Scatter Plots')
        st.pyplot(pair_plot.fig)

    # Box plots (for numeric columns)
    for column in numeric_columns:
        fig, ax = plt.subplots()
        sns.boxplot(x=data[column])
        plt.title(f'Box Plot of {column}')
        plt.xlabel(column)
        st.pyplot(fig)

    # Count plots (for categorical columns)
    categorical_columns = data.select_dtypes(include=['object']).columns
    for column in categorical_columns:
        fig, ax = plt.subplots()
        sns.countplot(data=data, x=column)
        plt.title(f'Count Plot of {column}')
        plt.xlabel(column)
        plt.ylabel('Count')
        plt.xticks(rotation=45)
        st.pyplot(fig)

    # Interactive scatter plot matrix (using Plotly)
    if len(numeric_columns) >= 2:
        fig = px.scatter_matrix(data, dimensions=numeric_columns, title='Interactive Scatter Plot Matrix')
        st.plotly_chart(fig)

    # Histograms for numeric columns
    for column in numeric_columns:
        fig, ax = plt.subplots()
        sns.histplot(data[column], kde=True)
        plt.title(f'Histogram of {column}')
        plt.xlabel(column)
        plt.ylabel('Frequency')
        st.pyplot(fig)
        
    return summary, data_types, missing_values, correlation_matrix

# Streamlit App
st.title("Automated EDA App")

# Define file_format variable
file_format = None

# Select dataset format
file_format = st.selectbox("Select a dataset format:", ["CSV", "XLS", "JSON", "TSV", "Inbuilt Datasets"])


if file_format == "Inbuilt Datasets":
    dataset_name = st.selectbox("Select an inbuilt dataset:", ["Iris", "Tips", "Titanic"])
    data = sns.load_dataset(dataset_name)
else:
    uploaded_file = st.file_uploader(f"Upload a {file_format} file", type=[file_format.lower()])
    if uploaded_file is not None:
        sheet_name = None
        if file_format in ["XLS", "XLSX"]:
            # Allow users to select a sheet from Excel files
            sheet_name = st.selectbox("Select a sheet (optional):", [""] + load_data(uploaded_file, file_format))
        data = load_data(uploaded_file, file_format, sheet_name=sheet_name)

if 'data' in locals():
    st.write("### Dataset Preview:")
    # Boolean to resize the dataframe, stored as a session state variable
    st.checkbox("Use container width", value=False, key="use_container_width")
    st.dataframe(data, use_container_width=st.session_state.use_container_width)

    st.write("### Automated EDA:")
    summary, data_types, missing_values, correlation_matrix = automated_eda(data)

    st.write("#### Summary Statistics:")
    st.write(summary)

    st.write("#### Data Types:")
    st.write(data_types)

    st.write("#### Missing Values:")
    st.write(missing_values)

    st.write("#### Correlation Matrix:")
    st.write(correlation_matrix)

    
    df_info = ['Automated EDA']
    sdbar = st.multiselect("EDA Options: ", df_info)
    # Button to perform EDA using Pandas Profiling
    if 'Automated EDA' in sdbar:
        datf = df
        st.write("Please Wait for Few Seconds.....")

        pr = df.profile_report(dark_mode=True)

        progress_bar = st.progress(0)
        status_text = st.empty()
        chart = st.line_chart(np.random.randn(10, 2))


        st_profile_report(pr)
        st.balloons()
